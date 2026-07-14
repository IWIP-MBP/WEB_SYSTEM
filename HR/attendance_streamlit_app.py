import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
import re

# 设置网页标题和布局
st.set_page_config(page_title="后勤三部考勤排休智能转换系统", layout="centered")

st.title("官方定制 📊 后勤三部考勤排休智能转换系统")
st.markdown("""
### 📅 月度通用操作指南：
1. **上传考勤明细**：点击下方框，可以**同时多选**并上传**上个月**和**当月**的系统原始出勤明细 Excel 文件（支持多选、跨月连贯追踪）。
2. **上传输出模板**：上传最新的空白 `输出模板.xlsx`。
3. **一键智能转换**：系统会自动穿透分析跨月流水、智能匹配出入园闭环轨迹、计算休息日加班并分级高亮，完成后提供标准 Excel 下载。
""")

def parse_ot_hours(val):
    """安全解析各种复杂的加班时长格式（如: 4h, 8h30m, 4.5, -, 0）并转换为浮点数"""
    if pd.isna(val): return 0.0
    val_str = str(val).lower().strip()
    if val_str in ['', '-', '0', '0.0', '0h']: return 0.0
    match = re.search(r'(?:(\d+(?:\.\d+)?)h)?\s*(?:(\d+(?:\.\d+)?)m)?', val_str)
    if match and (match.group(1) or match.group(2)):
        h = float(match.group(1)) if match.group(1) else 0.0
        m = float(match.group(2)) if match.group(2) else 0.0
        return h + (m / 60.0)
    try: return float(re.sub(r'[^\d.]', '', val_str))
    except: return 0.0

# --- 1. 网页端交互组件 ---
uploaded_logs = st.file_uploader("📂 步骤一：上传出勤明细 Excel 文件（可多选，请同时选中上月和本月流水）", type=["xlsx", "xls"], accept_multiple_files=True)
uploaded_template = st.file_uploader("📄 步骤二：上传您的『输出模板.xlsx』", type=["xlsx"])

if uploaded_logs and uploaded_template:
    if st.button("🚀 开始跨月深度追踪与转换"):
        with st.spinner("程序正在全局检索出勤轨迹并重构矩阵，请稍候..."):
            try:
                # --- 2. 多月份出勤流水进行全量合并与清洗 ---
                all_logs = []
                for file in uploaded_logs:
                    # 弹性行扫描：动态找寻包含“工号”和“日期”的核心标题行
                    df_file = pd.read_excel(file, header=None)
                    header_row = 0
                    for idx, row_vals in df_file.iterrows():
                        if "工号" in row_vals.values and "日期" in row_vals.values:
                            header_row = idx
                            break
                    df_cleaned = pd.read_excel(file, skiprows=header_row)
                    df_cleaned.columns = [str(c).strip() for c in df_cleaned.columns]
                    all_logs.append(df_cleaned)
                
                # 垂直合并
                df_all = pd.concat(all_logs, ignore_index=True)
                df_all['日期'] = pd.to_datetime(df_all['日期'].astype(str).str.strip(), errors='coerce')
                df_all = df_all.dropna(subset=['工号', '日期']).sort_values(by='日期') # 按时间严格排序
                df_all['工号'] = df_all['工号'].astype(str).str.split('.').str[0].str.strip()
                
                # 动态识别正在处理的最新目标月份（以流水中最新的时间作为本月）
                latest_date = df_all['日期'].max()
                current_year = latest_date.year
                current_month = latest_date.month
                
                # --- 3. 核心数据存储结构 ---
                attendance_matrix = {}   # {(工号, day): (状态文本, 颜色Fill对象)}
                emp_status_history = {}  # {工号: [(日期, "出园"/"入园")]} 用于还原完整出勤线
                
                for _, row in df_all.iterrows():
                    emp_id = row['工号']
                    date_val = row['日期']
                    shift_name = str(row.get('班次', '')).strip()
                    ot_hours = parse_ot_hours(row.get('加班时长', 0))
                    
                    if emp_id not in emp_status_history:
                        emp_status_history[emp_id] = []
                    
                    # 捕获关键状态变化
                    if "出园" in shift_name or "休假出园" in shift_name:
                        emp_status_history[emp_id].append((date_val, "出园"))
                    elif "入园" in shift_name:
                        emp_status_history[emp_id].append((date_val, "入园"))
                        
                    # 仅限本月的数据，写入 1-31 号的主体排休矩阵中
                    if date_val.year == current_year and date_val.month == current_month:
                        day = date_val.day
                        status_text = ""
                        cell_fill = None
                        
                        # 判断休息或休假班次
                        if "全天休" in shift_name or shift_name == "休息":
                            status_text = "全"
                            # 规则 2：平时休息不着色；但如果当天加班，按照时长进行分级标记
                            if ot_hours >= 8.0:
                                cell_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid") # 加班≥8小时：柔黄色
                            elif ot_hours >= 4.0:
                                cell_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid") # 加班≥4小时：柔绿色
                        elif "上午休" in shift_name or "上午休息" in shift_name:
                            status_text = "上"
                        elif "下午休" in shift_name or "下午休息" in shift_name:
                            status_text = "下"
                            
                        if status_text:
                            attendance_matrix[(emp_id, day)] = (status_text, cell_fill)

                # --- 4. 智能多维度备注分析（完美融合补充新规） ---
                computed_remarks = {} # {工号: 备注字符串}
                for emp_id, history in emp_status_history.items():
                    # 提取该员工在日志里的所有事件，按时间线划分前后匹配对
                    pairs = []
                    used_indices = set()
                    for idx, (d, event) in enumerate(history):
                        if idx in used_indices:
                            continue
                        if event == "出园":
                            out_date = d
                            in_date = None
                            # 搜寻此出园后紧接着的第一条入园记录
                            for j in range(idx + 1, len(history)):
                                if history[j][1] == "入园":
                                    in_date = history[j][0]
                                    used_indices.add(j)
                                    break
                                elif history[j][1] == "出园":
                                    break # 连续两个出园，说明上一次出园独立未归
                            pairs.append((out_date, in_date))
                        else:
                            # 孤立的入园记录（前面没有匹配到出园）
                            pairs.append((None, d))
                    
                    # 依据划分好的封闭区间，判断属于哪种备注场景
                    remark_segments = []
                    for out_date, in_date in pairs:
                        if out_date and in_date:
                            out_is_cur = (out_date.year == current_year and out_date.month == current_month)
                            in_is_cur = (in_date.year == current_year and in_date.month == current_month)
                            
                            if out_is_cur and in_is_cur:
                                # 💡 【补充新规】：本月开始休假，且本月入园的
                                remark_segments.append(f"{out_date.month}.{out_date.day}休假出园，{in_date.month}.{in_date.day}入园")
                            elif (not out_is_cur) and in_is_cur:
                                # 💡 【规则 3】：上个月已经休假，当月才入园的
                                remark_segments.append(f"{out_date.month}.{out_date.day}休假出园，{in_date.month}.{in_date.day}入园")
                            elif out_is_cur and (not in_is_cur):
                                # 本月出园，下月才进（如果在数据尾端）
                                remark_segments.append(f"{out_date.month}.{out_date.day}休假出园")
                        elif out_date and not in_date:
                            out_is_cur = (out_date.year == current_year and out_date.month == current_month)
                            if out_is_cur:
                                # 💡 【规则 3】：本月开始休假但没有入园的
                                remark_segments.append(f"{out_date.month}.{out_date.day}休假出园")
                        elif not out_date and in_date:
                            in_is_cur = (in_date.year == current_year and in_date.month == current_month)
                            if in_is_cur:
                                remark_segments.append(f"{in_date.month}.{in_date.day}入园")
                    
                    if remark_segments:
                        # 排除完全一致的重复行项并用“；”衔接
                        unique_segments = list(dict.fromkeys(remark_segments))
                        computed_remarks[emp_id] = "；".join(unique_segments)

                # --- 5. 安全载入并增量涂刷 Excel 模板 ---
                template_bytes = uploaded_template.read()
                wb = load_workbook(io.BytesIO(template_bytes))
                ws = wb["排休"] if "排休" in wb.sheetnames else wb.active
                
                # 动态捕捉列坐标，杜绝排版错位风险
                day_to_col = {}
                id_col, remark_col = None, None
                for r in range(1, 6):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if val:
                            val_str = str(val).strip()
                            if val_str.isdigit() and 1 <= int(val_str) <= 31:
                                day_to_col[int(val_str)] = c
                            if "工号" in val_str and id_col is None: id_col = c
                            if "备注" in val_str and remark_col is None: remark_col = c

                # 启动覆盖写入
                updated_count = 0
                for row in range(4, ws.max_row + 1):
                    cell_id = ws.cell(row=row, column=id_col).value if id_col else None
                    if cell_id is None: continue
                    emp_id = str(cell_id).split('.')[0].strip()
                    if not emp_id or emp_id in ['None', 'nan']: continue
                    
                    # 写入 1-31 号状态
                    for day in range(1, 32):
                        if day in day_to_col:
                            c_idx = day_to_col[day]
                            # 全面洗刷模板上遗留的过往历史数据与背景色
                            ws.cell(row=row, column=c_idx).value = None
                            ws.cell(row=row, column=c_idx).fill = PatternFill(fill_type=None)
                            
                            # 注入本月新计算出来的考勤
                            if (emp_id, day) in attendance_matrix:
                                status_text, cell_fill = attendance_matrix[(emp_id, day)]
                                ws.cell(row=row, column=c_idx).value = status_text
                                if cell_fill:
                                    ws.cell(row=row, column=c_idx).fill = cell_fill
                                    
                    # 写入跨月智能穿透备注
                    if emp_id in computed_remarks:
                        ws.cell(row=row, column=remark_col).value = computed_remarks[emp_id]
                    else:
                        ws.cell(row=row, column=remark_col).value = None # 清空旧备注
                        
                    updated_count += 1

                # --- 6. 将处理完成的表格转回二进制流供网页下载 ---
                out_buffer = io.BytesIO()
                wb.save(out_buffer)
                out_buffer.seek(0)
                
                st.success(f"✨ 转换大功告成！已成功重构【{current_year}年{current_month}月】调休排班数据，覆盖 {updated_count} 人。")
                st.download_button(
                    label="💾 点击下载更新后的排休表 Excel 附件",
                    data=out_buffer,
                    file_name=f"后勤三部_智能化生成排休表_{current_month}月.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"❌ 运行失败，可能原因：表格表头被修改或格式损坏。错误摘要: {str(e)}")
""
