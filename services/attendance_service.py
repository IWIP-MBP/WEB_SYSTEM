import io
import re
import logging
import copy
from datetime import datetime, date
from typing import List, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

logger = logging.getLogger("uvicorn.error")

def parse_ot_hours(val) -> float:
    """安全解析各种复杂的加班时长格式（如: 4h, 8h30m, 4.5, -, 0）并转换为浮点数"""
    if pd.isna(val):
        return 0.0
    val_str = str(val).lower().strip()
    if val_str in ['', '-', '0', '0.0', '0h']:
        return 0.0
    match = re.search(r'(?:(\d+(?:\.\d+)?)h)?\s*(?:(\d+(?:\.\d+)?)m)?', val_str)
    if match and (match.group(1) or match.group(2)):
        h = float(match.group(1)) if match.group(1) else 0.0
        m = float(match.group(2)) if match.group(2) else 0.0
        return h + (m / 60.0)
    try:
        return float(re.sub(r'[^\d.]', '', val_str))
    except Exception:
        return 0.0

def safe_insert_rows(ws, idx, amount=1):
    """
    安全地在工作表中插入行，并且将插入位置及下方的合并单元格范围向下平移，
    防止新插入的行被旧的合并单元格样式错误吞并。
    """
    ranges_to_shift = []
    other_ranges = []
    for r in list(ws.merged_cells.ranges):
        if r.min_row >= idx:
            ranges_to_shift.append(r)
        else:
            other_ranges.append(r)
            
    # 移除合并单元格以防止 insert_rows 造成样式污染
    ws.merged_cells.ranges.clear()
    for r in other_ranges:
        ws.merged_cells.add(r)
        
    ws.insert_rows(idx, amount)
    
    # 重新注册并平移合并单元格
    for r in ranges_to_shift:
        r.shift(row_shift=amount)
        ws.merged_cells.add(r)

def apply_row_style_and_border(ws, row_idx, ref_styles, ref_height):
    """
    对指定行设置细边框、居中对齐、行高，并应用参考行的字体和数字格式。
    """
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[row_idx].height = ref_height
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = center_alignment
        
        if col in ref_styles:
            style = ref_styles[col]
            if style['font']:
                cell.font = copy.copy(style['font'])
            if style['number_format']:
                cell.number_format = style['number_format']

def check_resigned_last_month(emp_id: str, current_year: int, current_month: int, emp_info: dict) -> bool:
    """检查员工是否在当前月之前就已经离职"""
    if emp_id not in emp_info:
        return False
    info = emp_info[emp_id]
    r_date = info['earliest_resign']
    if r_date is not None:
        if r_date.year < current_year or (r_date.year == current_year and r_date.month < current_month):
            return True
    return False

def extract_attendance_events(attendance_files: List[bytes], current_year: int) -> dict:
    """
    扫描所有出勤明细文件中的所有工作表，提取出入园及离职事件。
    返回: {
        emp_id: {
            'out': set([date, ...]),
            'in': set([date, ...]),
            'resign': set([date, ...]),
            'leave': set([date, ...])
        }
    }
    """
    emp_events = {}
    
    for file_bytes in attendance_files:
        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            logger.error(f"加载工作簿失败: {e}")
            continue
            
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Detect header row and 工号 column
            header_row_idx = None
            id_col = None
            date_col = None
            
            # Scan first 20 rows to find headers
            for r in range(1, min(25, ws.max_row + 1)):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                if "工号" in row_vals:
                    header_row_idx = r
                    id_col = row_vals.index("工号") + 1
                    if "日期" in row_vals:
                        date_col = row_vals.index("日期") + 1
                    break
            
            if header_row_idx is None:
                continue
                
            # If date_col is present, it's a Log Sheet (row-based)
            if date_col is not None:
                for r in range(header_row_idx + 1, ws.max_row + 1):
                    emp_id_val = ws.cell(row=r, column=id_col).value
                    if emp_id_val is None:
                        continue
                    emp_id = str(emp_id_val).split('.')[0].strip()
                    if not emp_id or emp_id in ['None', 'nan', '工号', '姓名', '序号', '备注']:
                        continue
                    
                    date_val = ws.cell(row=r, column=date_col).value
                    if date_val is None:
                        continue
                    
                    dt = None
                    if isinstance(date_val, datetime):
                        dt = date_val.date()
                    elif isinstance(date_val, date):
                        dt = date_val
                    elif isinstance(date_val, str):
                        try:
                            dt = pd.to_datetime(date_val.strip()).date()
                        except:
                            continue
                    else:
                        try:
                            dt = pd.to_datetime(str(date_val).strip()).date()
                        except:
                            continue
                            
                    if dt is None:
                        continue
                        
                    if emp_id not in emp_events:
                        emp_events[emp_id] = { 'out': set(), 'in': set(), 'resign': set(), 'leave': set() }
                        
                    # Scan all cells in this row
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if val is not None:
                            val_str = str(val)
                            if "出园回国" in val_str:
                                emp_events[emp_id]['out'].add(dt)
                            elif "出境入园" in val_str:
                                emp_events[emp_id]['in'].add(dt)
                            elif "已离职" in val_str or "离职" in val_str:
                                emp_events[emp_id]['resign'].add(dt)
                            elif "休假" in val_str or "请假" in val_str:
                                emp_events[emp_id]['leave'].add(dt)
                                
            # If date_col is not present, it's a Summary Sheet (column-based dates)
            else:
                # Find date columns in header row
                day_cols = {}
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=header_row_idx, column=c).value
                    if val:
                        val_str = str(val).strip()
                        # Match MM-DD
                        match = re.match(r'^(\d{1,2})-(\d{1,2})$', val_str)
                        if match:
                            m = int(match.group(1))
                            d = int(match.group(2))
                            try:
                                day_cols[c] = date(current_year, m, d)
                            except:
                                pass
                        else:
                            if isinstance(val, datetime):
                                day_cols[c] = val.date()
                            elif isinstance(val, date):
                                day_cols[c] = val
                            else:
                                try:
                                    day_cols[c] = pd.to_datetime(val_str).date()
                                except:
                                    pass
                                    
                # Scan employee rows
                for r in range(header_row_idx + 1, ws.max_row + 1):
                    emp_id_val = ws.cell(row=r, column=id_col).value
                    if emp_id_val is None:
                        continue
                    emp_id = str(emp_id_val).split('.')[0].strip()
                    if not emp_id or emp_id in ['None', 'nan', '工号', '姓名', '序号', '备注']:
                        continue
                        
                    if emp_id not in emp_events:
                        emp_events[emp_id] = { 'out': set(), 'in': set(), 'resign': set(), 'leave': set() }
                        
                    for c, dt in day_cols.items():
                        cell_val = ws.cell(row=r, column=c).value
                        if cell_val is not None:
                            val_str = str(cell_val)
                            if "出园回国" in val_str:
                                emp_events[emp_id]['out'].add(dt)
                            elif "出境入园" in val_str:
                                emp_events[emp_id]['in'].add(dt)
                            elif "已离职" in val_str or "离职" in val_str:
                                emp_events[emp_id]['resign'].add(dt)
                            elif "休假" in val_str or "请假" in val_str:
                                emp_events[emp_id]['leave'].add(dt)
                                
    return emp_events

def compile_emp_info(emp_events) -> dict:
    """计算每个员工的出入园、离职的极值，用于生成假期区间和备注"""
    emp_info = {}
    for emp_id, ev in emp_events.items():
        out_dates = sorted(list(ev['out']))
        in_dates = sorted(list(ev['in']))
        resign_dates = sorted(list(ev['resign']))
        
        earliest_out = out_dates[0] if out_dates else None
        latest_in = in_dates[-1] if in_dates else None
        earliest_resign = resign_dates[0] if resign_dates else None
        
        emp_info[emp_id] = {
            'earliest_out': earliest_out,
            'latest_in': latest_in,
            'earliest_resign': earliest_resign,
            'leave_dates': ev['leave'],
            'out_dates': ev['out'],
            'in_dates': ev['in'],
            'resign_dates': ev['resign']
        }
    return emp_info

def check_inactive_or_leave(emp_id: str, d: date, emp_info: dict) -> bool:
    """检查特定工号在某天是否属于休假/非在职状态"""
    if emp_id not in emp_info:
        return False
    info = emp_info[emp_id]
    
    # 1. 离职状态检查
    if info['earliest_resign'] is not None and d >= info['earliest_resign']:
        return True
        
    # 2. 出入园休假区间检查
    out_date = info['earliest_out']
    in_date = info['latest_in']
    if out_date is not None and in_date is not None:
        if out_date <= d <= in_date:
            return True
    elif out_date is None and in_date is not None:
        if d <= in_date:
            return True
    elif out_date is not None and in_date is None:
        if d >= out_date:
            return True
            
    # 3. 单日明细中明确标记“休假”、“请假”、“出园”或“入园”的检查
    if (d in info['leave_dates'] or 
        d in info['out_dates'] or 
        d in info['in_dates'] or 
        d in info['resign_dates']):
        return True
        
    return False

def convert_attendance(attendance_files: List[bytes], template_bytes: bytes) -> Tuple[bytes, int, int, int]:
    """
    处理多月份的出勤明细 Excel 文件并按模板更新排休。
    返回: (更新后的Excel字节流, 最新目标年, 最新目标月, 更新的人数)
    """
    if not attendance_files:
        raise ValueError("未上传任何出勤明细文件")
    if not template_bytes:
        raise ValueError("未上传排休模板文件")

    all_logs = []
    for file_bytes in attendance_files:
        file_stream = io.BytesIO(file_bytes)
        xls = pd.ExcelFile(file_stream)
        target_sheet = None
        header_row = 0
        for sheet_name in xls.sheet_names:
            df_file = xls.parse(sheet_name, header=None)
            found = False
            for idx, row_vals in df_file.iterrows():
                if "工号" in row_vals.values and "日期" in row_vals.values:
                    header_row = idx
                    target_sheet = sheet_name
                    found = True
                    break
            if found:
                break
        
        if target_sheet is None:
            target_sheet = xls.sheet_names[0]
            df_file = xls.parse(target_sheet, header=None)
            header_row = 0
            for idx, row_vals in df_file.iterrows():
                if "工号" in row_vals.values and "日期" in row_vals.values:
                    header_row = idx
                    break
        
        df_cleaned = xls.parse(target_sheet, skiprows=header_row)
        df_cleaned.columns = [str(c).strip() for c in df_cleaned.columns]
        logger.info(f"成功解析工作表 '{target_sheet}'，表头行位于 {header_row}。包含列: {list(df_cleaned.columns)}")
        all_logs.append(df_cleaned)
    
    # 垂直合并
    df_all = pd.concat(all_logs, ignore_index=True)
    df_all['日期'] = pd.to_datetime(df_all['日期'].astype(str).str.strip(), errors='coerce')
    df_all = df_all.dropna(subset=['工号', '日期']).sort_values(by='日期') # 按时间严格排序
    df_all['工号'] = df_all['工号'].astype(str).str.split('.').str[0].str.strip()
    
    unique_ids_logs = list(df_all['工号'].unique())
    logger.info(f"合并后的出勤记录共计 {len(df_all)} 条。包含的工号数: {len(unique_ids_logs)}。")
    
    # 动态识别正在处理的最新目标月份（以流水中最新的时间作为本月）
    latest_date = df_all['日期'].max()
    if pd.isna(latest_date):
        raise ValueError("未能从出勤明细中提取到有效的『日期』数据，请检查文件格式")
        
    current_year = latest_date.year
    current_month = latest_date.month
    
    # --- 提取出入园、离职等事件及备注信息 ---
    emp_events = extract_attendance_events(attendance_files, current_year)
    emp_info = compile_emp_info(emp_events)
    
    # --- 3. 核心数据存储结构 ---
    attendance_matrix = {}   # {(工号, day): (状态文本, 颜色Fill对象)}
    
    for _, row in df_all.iterrows():
        emp_id = row['工号']
        date_val = row['日期']
        if pd.isna(date_val):
            continue
        shift_name = str(row.get('班次', '')).strip()
        ot_hours = parse_ot_hours(row.get('加班时长', 0))
            
        # 仅限本月的数据，写入 1-31 号的主体排休矩阵中
        if date_val.year == current_year and date_val.month == current_month:
            day = date_val.day
            status_text = ""
            cell_fill = None
            
            # 判断休息或休假班次
            if "全天休" in shift_name or shift_name == "休息":
                if ot_hours >= 4.0:
                    status_text = ""
                    if ot_hours >= 8.0:
                        cell_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 加班≥8小时：明黄色
                    else:
                        cell_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid") # 加班≥4小时：中绿色
                else:
                    status_text = "全"
                    cell_fill = None
            elif "上午休" in shift_name or "上午休息" in shift_name:
                status_text = "上"
            elif "下午休" in shift_name or "下午休息" in shift_name:
                status_text = "下"
                
            if status_text or cell_fill:
                attendance_matrix[(emp_id, day)] = (status_text, cell_fill)

    # --- 4. 智能多维度备注分析 ---
    computed_remarks = {}
    for emp_id, info in emp_info.items():
        earliest_out = info['earliest_out']
        latest_in = info['latest_in']
        earliest_resign = info['earliest_resign']
        
        vacation_part = ""
        out_in_current = (earliest_out is not None and earliest_out.year == current_year and earliest_out.month == current_month)
        in_in_current = (latest_in is not None and latest_in.year == current_year and latest_in.month == current_month)
        
        # 规则 3 & 5：上个月入园的不显示在备注中，日期格式更改为 X.X
        if in_in_current:
            if earliest_out is not None:
                vacation_part = f"{earliest_out.month}.{earliest_out.day}出园回国{latest_in.month}.{latest_in.day}入园"
            else:
                vacation_part = f"{latest_in.month}.{latest_in.day}入园"
        else:
            if earliest_out is not None and out_in_current:
                vacation_part = f"{earliest_out.month}.{earliest_out.day}出园回国"
                
        resign_part = ""
        if earliest_resign is not None:
            resign_part = f"{earliest_resign.month}.{earliest_resign.day}离职"
            
        parts = []
        if vacation_part:
            parts.append(vacation_part)
        if resign_part:
            parts.append(resign_part)
            
        if parts:
            computed_remarks[emp_id] = "；".join(parts)

    # --- 5. 安全载入并增量涂刷 Excel 模板 ---
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb["排休"] if "排休" in wb.sheetnames else wb.active
    
    # 自动更正模板中的月份
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).strip()
                is_month_cell = False
                if "月份" in val_str:
                    is_month_cell = True
                elif re.search(r'\d{4}[-/年]\d{1,2}', val_str):
                    is_month_cell = True
                elif isinstance(val, (datetime, date)):
                    is_month_cell = True
                
                if is_month_cell:
                    if "：" in val_str:
                        ws.cell(row=r, column=c).value = f"月份：{current_year}年{current_month}月"
                    elif ":" in val_str:
                        ws.cell(row=r, column=c).value = f"月份:{current_year}年{current_month}月"
                    else:
                        ws.cell(row=r, column=c).value = f"月份：{current_year}年{current_month}月"
                    logger.info(f"已更正月份单元格 (行 {r}, 列 {c}) 为：{ws.cell(row=r, column=c).value}")

    # 自动增加或删除 31日列，并更正周几
    date_row_idx = None
    day_1_col = None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                try:
                    val_int = int(str(val).strip())
                    if val_int == 1:
                        # 检查下一列以确认是日期行
                        next_val = ws.cell(row=r, column=c+1).value
                        if next_val is not None:
                            try:
                                next_val_int = int(str(next_val).strip())
                                if next_val_int == 2:
                                    date_row_idx = r
                                    day_1_col = c
                                    break
                            except ValueError:
                                pass
                except ValueError:
                    pass
        if date_row_idx is not None:
            break

    if date_row_idx is not None and day_1_col is not None:
        template_days = 0
        c = day_1_col
        while True:
            val = ws.cell(row=date_row_idx, column=c).value
            if val is not None:
                try:
                    val_int = int(str(val).strip())
                    if val_int == template_days + 1:
                        template_days = val_int
                        c += 1
                        continue
                except ValueError:
                    pass
            break
            
        import calendar
        target_days = calendar.monthrange(current_year, current_month)[1]
        
        if template_days > 0 and template_days != target_days:
            if template_days > target_days:
                # 删除多余列
                start_del_col = day_1_col + target_days
                num_to_delete = template_days - target_days
                
                # 手动平移/调整合并单元格
                from openpyxl.worksheet.cell_range import CellRange
                ranges = list(ws.merged_cells.ranges)
                ws.merged_cells.ranges.clear()
                for r in ranges:
                    min_col, min_row, max_col, max_row = r.min_col, r.min_row, r.max_col, r.max_row
                    if min_col >= start_del_col and max_col < start_del_col + num_to_delete:
                        continue
                    if min_col >= start_del_col + num_to_delete:
                        min_col -= num_to_delete
                        max_col -= num_to_delete
                    elif min_col >= start_del_col:
                        min_col = start_del_col
                        max_col -= num_to_delete
                    elif max_col >= start_del_col:
                        max_col -= num_to_delete
                    if min_col <= max_col:
                        new_r = CellRange(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
                        ws.merged_cells.add(new_r)
                
                ws.delete_cols(start_del_col, num_to_delete)
                logger.info(f"已自动删除模板中多余的 {num_to_delete} 天日期列（从第 {start_del_col} 列开始）")
                
            else:
                # 插入缺少列
                start_ins_col = day_1_col + template_days
                num_to_insert = target_days - template_days
                
                # 计算原有日期列的最小宽度作为新插入列的宽度（保证最窄/和原有日期列一致）
                day_widths = []
                for day_c in range(day_1_col, day_1_col + template_days):
                    import openpyxl
                    col_letter = openpyxl.utils.get_column_letter(day_c)
                    w = ws.column_dimensions[col_letter].width
                    if w is not None and w > 0:
                        day_widths.append(w)
                target_width = min(day_widths) if day_widths else 4.625

                # 手动平移/调整合并单元格
                from openpyxl.worksheet.cell_range import CellRange
                ranges = list(ws.merged_cells.ranges)
                ws.merged_cells.ranges.clear()
                for r in ranges:
                    min_col, min_row, max_col, max_row = r.min_col, r.min_row, r.max_col, r.max_row
                    if min_col >= start_ins_col:
                        min_col += num_to_insert
                        max_col += num_to_insert
                    elif max_col >= start_ins_col:
                        max_col += num_to_insert
                    new_r = CellRange(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
                    ws.merged_cells.add(new_r)
                    
                ws.insert_cols(start_ins_col, num_to_insert)
                logger.info(f"已自动插入模板中缺少的 {num_to_insert} 天日期列（从第 {start_ins_col} 列开始）")
                
                # 复制样式到新插入列
                for i in range(num_to_insert):
                    new_col = start_ins_col + i
                    new_day = template_days + 1 + i
                    
                    # 设置新插入的日期列宽度为原有日期列的最小宽度（最窄）
                    import openpyxl
                    dst_col_letter = openpyxl.utils.get_column_letter(new_col)
                    ws.column_dimensions[dst_col_letter].width = target_width
                    
                    # 复制单元格样式
                    src_col = start_ins_col - 1
                    for row in range(1, ws.max_row + 1):
                        src_cell = ws.cell(row=row, column=src_col)
                        dst_cell = ws.cell(row=row, column=new_col)
                        if src_cell.has_style:
                            dst_cell.font = copy.copy(src_cell.font)
                            dst_cell.border = copy.copy(src_cell.border)
                            dst_cell.fill = copy.copy(src_cell.fill)
                            dst_cell.number_format = src_cell.number_format
                            dst_cell.alignment = copy.copy(src_cell.alignment)
                            
                    # 填充日期数字
                    ws.cell(row=date_row_idx, column=new_col).value = new_day

        # 自动更正周几对应的行
        weekdays_zh = ["一", "二", "三", "四", "五", "六", "日"]
        for d in range(1, target_days + 1):
            col = day_1_col + d - 1
            dt = date(current_year, current_month, d)
            wk_name = weekdays_zh[dt.weekday()]
            ws.cell(row=date_row_idx + 1, column=col).value = wk_name
        logger.info(f"已自动更正所有日期的周几显示（行 {date_row_idx + 1}）")
    else:
        import calendar
        target_days = calendar.monthrange(current_year, current_month)[1]

    # 动态捕捉列坐标，杜绝排版错位风险
    day_to_col = {}
    id_col, remark_col = None, None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                val_str = str(val).strip()
                if val_str.isdigit() and 1 <= int(val_str) <= 31:
                    day_to_col[int(val_str)] = c
                if "工号" in val_str and id_col is None:
                    id_col = c
                if "备注" in val_str and remark_col is None:
                    remark_col = c

    if id_col is None:
        raise ValueError("排休模板中未找到包含'工号'的标识列")

    logger.info(f"模板解析完成：工号列为第 {id_col} 列，备注列为第 {remark_col} 列，已匹配天数对应的列数: {len(day_to_col)}")

    # 复制第 6 行的样式以作为新增行的样板，防止插入行后样式丢失
    ref_styles = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=6, column=col)
        ref_styles[col] = {
            'font': copy.copy(cell.font) if cell.font else None,
            'number_format': cell.number_format
        }
    ref_height = ws.row_dimensions[6].height or 22.0

    # 检查模板中是否已包含工号
    rows_to_check = []
    for row in range(6, ws.max_row + 1):
        cell_id = ws.cell(row=row, column=id_col).value
        if cell_id is not None:
            emp_id = str(cell_id).split('.')[0].strip()
            if emp_id and emp_id not in ['None', 'nan', '工号', '姓名', '序号', '备注']:
                rows_to_check.append((row, emp_id))
                
    # 规则 4：上个月已离职的员工不需要体现。从下往上删除模板中的对应行。
    rows_to_delete = []
    for row, emp_id in rows_to_check:
        if check_resigned_last_month(emp_id, current_year, current_month, emp_info):
            rows_to_delete.append((row, emp_id))
            
    for row, emp_id in sorted(rows_to_delete, key=lambda x: x[0], reverse=True):
        ws.delete_rows(row)
        logger.info(f"由于上月已离职，已从模板中删除工号 {emp_id} 所在的第 {row} 行。")

    # 重新构建模板中剩下的有效行信息
    template_row_ids = []
    for row in range(6, ws.max_row + 1):
        cell_id = ws.cell(row=row, column=id_col).value
        if cell_id is not None:
            emp_id = str(cell_id).split('.')[0].strip()
            if emp_id and emp_id not in ['None', 'nan', '工号', '姓名', '序号', '备注']:
                template_row_ids.append((row, emp_id))
                
    # 如果模板为空，自动填充员工
    if not template_row_ids:
        logger.info("检测到模板内无有效工号，启动自动填充员工名单逻辑...")
        unique_emps = df_all.drop_duplicates(subset=['工号']).sort_values(by='工号')
        row_idx = 6
        for _, emp_row in unique_emps.iterrows():
            emp_id = str(emp_row['工号']).strip()
            emp_name = str(emp_row.get('姓名', '')).strip()
            
            # 跳过上月已离职的员工
            if check_resigned_last_month(emp_id, current_year, current_month, emp_info):
                continue
                
            if row_idx >= 30:
                safe_insert_rows(ws, row_idx, 1)
                
            apply_row_style_and_border(ws, row_idx, ref_styles, ref_height)
            
            ws.cell(row=row_idx, column=1).value = row_idx - 5
            ws.cell(row=row_idx, column=id_col).value = emp_id
            ws.cell(row=row_idx, column=3).value = emp_name
            
            template_row_ids.append((row_idx, emp_id))
            row_idx += 1
            
        logger.info(f"已自动向模板追加写入 {len(template_row_ids)} 名员工信息。")
    else:
        logger.info(f"模板中已含有已填充的工号行（共 {len(template_row_ids)} 行）。")
        # 确保模板已有的行也统一设置边框和居中格式
        for row, _ in template_row_ids:
            apply_row_style_and_border(ws, row, ref_styles, ref_height)

    # 规则 1：强制解开员工数据行（第6行至最大员工行）的合并单元格，禁止结果数据中包含任何合并单元格
    if template_row_ids:
        max_emp_row = max(r for r, _ in template_row_ids)
        merged_ranges = list(ws.merged_cells.ranges)
        for r in merged_ranges:
            if r.max_row >= 6 and r.min_row <= max_emp_row:
                try:
                    ws.unmerge_cells(r.coord)
                    logger.info(f"已解开数据区内的合并单元格: {r.coord}")
                except Exception as unmerge_err:
                    logger.warning(f"解开合并单元格 {r.coord} 失败: {unmerge_err}")

    # 执行覆盖写入状态及备注
    updated_count = 0
    for row, emp_id in template_row_ids:
        # 写入本月所有日期状态
        for day in range(1, target_days + 1):
            if day in day_to_col:
                c_idx = day_to_col[day]
                # 全面洗刷模板上遗留的过往历史数据与背景色
                ws.cell(row=row, column=c_idx).value = None
                ws.cell(row=row, column=c_idx).fill = PatternFill(fill_type=None)
                
                # Check if this day is a leave/inactive day
                d = date(current_year, current_month, day)
                if check_inactive_or_leave(emp_id, d, emp_info):
                    # It is a leave or inactive day. Leave it completely empty!
                    continue
                
                # 注入本月新计算出来的考勤
                if (emp_id, day) in attendance_matrix:
                    status_text, cell_fill = attendance_matrix[(emp_id, day)]
                    ws.cell(row=row, column=c_idx).value = status_text
                    if cell_fill:
                        ws.cell(row=row, column=c_idx).fill = cell_fill
                        
        # 写入跨月智能穿透备注并设置自动换行
        if remark_col is not None:
            remark_cell = ws.cell(row=row, column=remark_col)
            if emp_id in computed_remarks:
                remark_cell.value = computed_remarks[emp_id]
            else:
                remark_cell.value = None # 清空旧备注
            
            # 设置自动换行，同时保留原有的对齐属性（如居中对齐）
            orig_align = remark_cell.alignment
            if orig_align:
                remark_cell.alignment = Alignment(
                    horizontal=orig_align.horizontal,
                    vertical=orig_align.vertical,
                    text_rotation=orig_align.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=orig_align.shrink_to_fit,
                    indent=orig_align.indent
                )
            else:
                remark_cell.alignment = Alignment(wrap_text=True, vertical='center')
            
        updated_count += 1

    logger.info(f"模板状态与数据写入完成。写入行数: {updated_count}")

    # --- 6. 将处理完成的表格转回二进制流 ---
    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer.getvalue(), current_year, current_month, updated_count
